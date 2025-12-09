"""
制热模式测试脚本
根据test_hot.xlsx中的测试数据运行测试

时间压缩规则：
- 测试用例每行表示1分钟（60秒系统时间）
- 测试过程中每行压缩为10秒实际时间
- 即：10秒测试时间 = 60秒系统时间
- 时间压缩比：TIME_SCALE = 6
"""

import os
import sys
import time
import threading
from datetime import datetime, timedelta
from decimal import Decimal
from openpyxl import load_workbook

# 设置 Django 环境
sys.path.append(os.path.join(os.path.dirname(__file__), "backend"))
os.environ.setdefault("DJANGO_SETTINGS_MODULE", "hotel_ac.settings")

import django
django.setup()

from ac_system.models import Room, ACState, ACDetailRecord, Customer, AccommodationOrder
from ac_system.scheduler import scheduler, ServiceObject, WaitingObject
from django.utils import timezone

# ============================================================
# 测试配置
# ============================================================

# 时间压缩比：10秒测试时间 = 60秒系统时间
TIME_SCALE = 6  # 系统时间 = 测试时间 * TIME_SCALE
TEST_INTERVAL = 10  # 每行测试数据间隔10秒

# 房间初始温度配置
INITIAL_TEMPS = {
    "101": 10.0,  # 房间一 10℃
    "102": 15.0,  # 房间二 15℃
    "103": 18.0,  # 房间三 18℃
    "104": 12.0,  # 房间四 12℃
    "105": 14.0,  # 房间五 14℃
}

# 制热模式默认温度
DEFAULT_HEATING_TEMP = 25

# 风速映射
FAN_SPEED_MAP = {
    "高": "high",
    "中": "medium",
    "低": "low",
}

# ============================================================
# 测试数据解析
# ============================================================

def parse_test_data(filepath):
    """
    解析测试数据Excel文件
    返回：[(时间, {房间号: 操作}), ...]
    """
    wb = load_workbook(filepath)
    ws = wb.active
    
    rows = list(ws.iter_rows(values_only=True))
    
    # 跳过前两行（标题行）
    data_rows = rows[2:]
    
    test_actions = []
    
    for row in data_rows:
        if row[0] is None or row[0] == '费用小计':
            continue
            
        time_min = row[0]
        if not isinstance(time_min, (int, float)):
            continue
        
        time_min = int(time_min)
        actions = {}
        
        # 解析每个房间的操作（列1-5对应房间1-5）
        for room_idx in range(5):
            room_id = f"10{room_idx + 1}"
            cell_value = row[room_idx + 1]
            
            if cell_value is not None:
                action = parse_action(cell_value)
                if action:
                    actions[room_id] = action
        
        if actions or time_min == 0:
            test_actions.append((time_min, actions))
    
    return test_actions


def parse_action(cell_value):
    """
    解析单元格中的操作指令
    可能的格式：
    - "开机"
    - "关机"
    - "高"/"中"/"低" (调风速)
    - 数字 (调温度)
    - "24，高" 或 "21，高" (同时调温度和风速)
    """
    if cell_value is None:
        return None
    
    cell_str = str(cell_value).strip()
    
    if cell_str == "开机":
        return {"type": "power_on"}
    elif cell_str == "关机":
        return {"type": "power_off"}
    elif cell_str in FAN_SPEED_MAP:
        return {"type": "change_speed", "fan_speed": FAN_SPEED_MAP[cell_str]}
    elif cell_str.replace(".", "").isdigit():
        return {"type": "change_temp", "target_temp": float(cell_str)}
    elif "，" in cell_str or "," in cell_str:
        # 同时调温度和风速，如 "21，高"
        parts = cell_str.replace("，", ",").split(",")
        result = {"type": "change_both"}
        for part in parts:
            part = part.strip()
            if part in FAN_SPEED_MAP:
                result["fan_speed"] = FAN_SPEED_MAP[part]
            elif part.replace(".", "").isdigit():
                result["target_temp"] = float(part)
        return result
    elif cell_str == "中央空调启动":
        return {"type": "system_start"}
    elif "检查程序" in cell_str or "设置" in cell_str:
        return None  # 忽略说明文字
    
    return None


# ============================================================
# 测试执行
# ============================================================

class HeatingTest:
    def __init__(self):
        self.room_ids = ["101", "102", "103", "104", "105"]
        self.room_states = {}  # 记录每个房间的当前设置
        self.test_start_time = None
        self.log_file = None
        
    def setup(self):
        """初始化测试环境"""
        print("=" * 60)
        print("制热模式测试 - 初始化")
        print("=" * 60)
        
        # 清理旧数据
        print("清理旧的测试数据...")
        ACDetailRecord.objects.filter(room_id__in=self.room_ids).delete()
        
        # 确保房间存在
        for room_id in self.room_ids:
            room, created = Room.objects.get_or_create(
                room_id=room_id,
                defaults={"room_type": "standard", "price_per_day": 400}
            )
            if created:
                print(f"  创建房间 {room_id}")
        
        # 创建测试顾客和订单
        for room_id in self.room_ids:
            customer, _ = Customer.objects.get_or_create(
                id_card=f"1234567890{room_id}",
                defaults={"name": f"测试顾客{room_id}", "phone": "13800138000"}
            )
            
            # 检查是否已有活跃订单
            active_order = AccommodationOrder.objects.filter(
                room_id=room_id, status="active"
            ).first()
            
            if not active_order:
                room = Room.objects.get(room_id=room_id)
                AccommodationOrder.objects.create(
                    customer=customer,
                    room=room,
                    status="active"
                )
                print(f"  创建房间 {room_id} 的入住订单")
        
        # 初始化调度器中的房间状态
        for room_id in self.room_ids:
            scheduler.init_room(room_id)
            # 设置初始温度
            initial_temp = INITIAL_TEMPS.get(room_id, 15.0)
            scheduler.service_manager.room_states[room_id]["current_temp"] = initial_temp
            scheduler.service_manager.room_states[room_id]["initial_temp"] = initial_temp
            scheduler.service_manager.room_states[room_id]["mode"] = "heating"
            
            self.room_states[room_id] = {
                "target_temp": DEFAULT_HEATING_TEMP,
                "fan_speed": "medium",
                "is_on": False,
            }
            print(f"  房间 {room_id} 初始温度: {initial_temp}°C")
        
        print()
    
    def execute_action(self, room_id, action, current_time):
        """执行单个操作"""
        action_type = action.get("type")
        
        if action_type == "power_on":
            target_temp = self.room_states[room_id]["target_temp"]
            fan_speed = self.room_states[room_id]["fan_speed"]
            
            scheduler.submit_request(room_id, {
                "action": "power_on",
                "target_temp": target_temp,
                "fan_speed": fan_speed,
                "mode": "heating",
            })
            self.room_states[room_id]["is_on"] = True
            print(f"    房间 {room_id}: 开机 (目标温度={target_temp}°C, 风速={fan_speed})")
            
        elif action_type == "power_off":
            scheduler.submit_request(room_id, {"action": "power_off"})
            self.room_states[room_id]["is_on"] = False
            print(f"    房间 {room_id}: 关机")
            
        elif action_type == "change_temp":
            target_temp = action.get("target_temp")
            self.room_states[room_id]["target_temp"] = target_temp
            # 只有在开机状态下才发送调温请求
            if self.room_states[room_id]["is_on"]:
                scheduler.submit_request(room_id, {
                    "action": "change_temp",
                    "target_temp": target_temp,
                    "mode": "heating",
                })
                print(f"    房间 {room_id}: 调温 -> {target_temp}°C")
            else:
                print(f"    房间 {room_id}: 设置目标温度 -> {target_temp}°C (未开机)")
            
        elif action_type == "change_speed":
            fan_speed = action.get("fan_speed")
            self.room_states[room_id]["fan_speed"] = fan_speed
            # 只有在开机状态下才发送调风速请求
            if self.room_states[room_id]["is_on"]:
                scheduler.submit_request(room_id, {
                    "action": "change_speed",
                    "fan_speed": fan_speed,
                })
                print(f"    房间 {room_id}: 调风速 -> {fan_speed}")
            else:
                print(f"    房间 {room_id}: 设置风速 -> {fan_speed} (未开机)")
            
        elif action_type == "change_both":
            target_temp = action.get("target_temp")
            fan_speed = action.get("fan_speed")
            
            if target_temp:
                self.room_states[room_id]["target_temp"] = target_temp
                if self.room_states[room_id]["is_on"]:
                    scheduler.submit_request(room_id, {
                        "action": "change_temp",
                        "target_temp": target_temp,
                        "mode": "heating",
                    })
            
            if fan_speed:
                self.room_states[room_id]["fan_speed"] = fan_speed
                if self.room_states[room_id]["is_on"]:
                    scheduler.submit_request(room_id, {
                        "action": "change_speed",
                        "fan_speed": fan_speed,
                    })
                scheduler.submit_request(room_id, {
                    "action": "change_speed",
                    "fan_speed": fan_speed,
                })
            
            print(f"    房间 {room_id}: 调温={target_temp}°C, 调风速={fan_speed}")
    
    def print_status(self, time_min):
        """打印当前所有房间状态"""
        print(f"\n  [状态] 时间={time_min}分钟")
        print("  " + "-" * 80)
        print(f"  {'房间':<8} {'状态':<10} {'当前温度':<12} {'目标温度':<12} {'风速':<10} {'费用':<10}")
        print("  " + "-" * 80)
        
        for room_id in self.room_ids:
            state = scheduler.get_room_state(room_id)
            status = state.get("status", "off")
            current_temp = state.get("current_temp", 0)
            target_temp = state.get("target_temp", 0)
            fan_speed = state.get("fan_speed", "medium")
            cost = state.get("cost", 0)
            
            # 标记队列位置
            if room_id in scheduler.service_queue:
                status = f"{status}[服务]"
            elif room_id in scheduler.wait_queue:
                wobj = scheduler.wait_queue[room_id]
                remaining = wobj.get_remaining_wait_time()
                status = f"{status}[等待{remaining:.0f}s]"
            
            print(f"  {room_id:<8} {status:<10} {current_temp:<12.1f} {target_temp:<12.1f} {fan_speed:<10} {cost:<10.2f}")
        
        print("  " + "-" * 80)
        print(f"  服务队列: {list(scheduler.service_queue.keys())}")
        print(f"  等待队列: {list(scheduler.wait_queue.keys())}")
    
    def run_test(self, test_data):
        """运行测试"""
        print("\n" + "=" * 60)
        print("开始执行测试")
        print(f"时间压缩比: {TIME_SCALE}x (10秒测试时间 = 60秒系统时间)")
        print("=" * 60)
        
        self.test_start_time = time.time()  # 提前一点时间，避免调度器启动延迟影响
        
        # 启动调度器
        scheduler.start()
        print("调度器已启动\n")
        
        current_time_idx = 0
        
        for time_min, actions in test_data:
            # 等待到达指定时间点
            target_test_time = time_min * TEST_INTERVAL
            current_test_time = time.time() - self.test_start_time
            
            if target_test_time > current_test_time:
                wait_time = target_test_time - current_test_time
                print(f"\n等待 {wait_time:.1f} 秒到达时间点 {time_min} 分钟...")
                time.sleep(wait_time)
            
            print(f"\n{'='*60}")
            print(f"时间点: {time_min} 分钟 (测试时间: {time.time() - self.test_start_time:.1f}秒)")
            print(f"{'='*60}")
            
            if time_min == 0:
                print("  系统启动，设置制热模式")
                continue
            
            # 执行该时间点的所有操作
            if actions:
                print("  执行操作:")
                for room_id, action in actions.items():
                    self.execute_action(room_id, action, time_min)
            else:
                print("  (无操作)")
            
            # 打印状态
            self.print_status(time_min)
        
        # 测试结束，打印最终状态
        print("\n" + "=" * 60)
        print("测试完成 - 最终状态")
        print("=" * 60)
        self.print_final_report()
        
        # 停止调度器
        scheduler.stop()
        print("\n调度器已停止")
    
    def print_final_report(self):
        """打印最终报告"""
        print("\n费用汇总（从详单记录统计）:")
        print("-" * 60)
        
        total_cost = Decimal("0.00")
        total_energy = 0.0
        
        for room_id in self.room_ids:
            # 从详单记录统计费用
            records = ACDetailRecord.objects.filter(room_id=room_id)
            room_cost = sum(r.cost for r in records)
            room_energy = sum(r.energy_consumed for r in records)
            total_cost += room_cost
            total_energy += room_energy
            print(f"  房间 {room_id}: 费用={room_cost:.2f}元, 能耗={room_energy:.2f}度")
        
        print("-" * 60)
        print(f"  总费用: {total_cost:.2f}元, 总能耗: {total_energy:.2f}度")
        
        # 打印详单记录
        print("\n详单记录:")
        print("-" * 100)
        print(f"  {'房间':<8} {'开始时间':<12} {'结束时间':<12} {'起始温度':<10} {'结束温度':<10} {'风速':<8} {'费用':<10}")
        print("-" * 100)
        records = ACDetailRecord.objects.filter(room_id__in=self.room_ids).order_by("start_time")
        for record in records:
            end_time = record.end_time.strftime("%H:%M:%S") if record.end_time else "进行中"
            end_temp = f"{record.end_temp:.1f}" if record.end_temp else "-"
            print(f"  {record.room_id:<8} {record.start_time.strftime('%H:%M:%S'):<12} {end_time:<12} "
                  f"{record.start_temp:<10.1f} {end_temp:<10} {record.fan_speed:<8} {record.cost:.2f}元")


def main():
    """主函数"""
    print("=" * 60)
    print("制热模式测试脚本")
    print(f"启动时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print("=" * 60)
    
    # 修改调度器配置以适应加速测试
    print("\n应用时间压缩配置...")
    import config
    config.TIME_SCALE = TIME_SCALE
    
    try:
        # 解析测试数据
        print("解析测试数据...")
        test_data = parse_test_data("test_hot.xlsx")
        print(f"共解析 {len(test_data)} 个时间点的测试数据\n")
        
        # 打印测试数据预览
        print("测试数据预览:")
        for time_min, actions in test_data[:10]:
            if actions:
                print(f"  {time_min}分钟: {actions}")
        if len(test_data) > 10:
            print(f"  ... 共 {len(test_data)} 个时间点")
        print()
        
        # 创建测试实例并运行
        test = HeatingTest()
        test.setup()
        test.run_test(test_data)
        
    finally:
        # 恢复原始配置
        print("\n恢复调度器原始配置...")
        config.TIME_SCALE = 1
    
    print("\n测试脚本执行完毕")


if __name__ == "__main__":
    main()
