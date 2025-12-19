#!/usr/bin/env python3
"""
åˆ¶çƒ­æ¨¡å¼APIæµ‹è¯•é©±åŠ¨å™¨
é€šè¿‡HTTPæ¥å£æ¨¡æ‹Ÿç”¨æˆ·æ“ä½œï¼Œè®©å‰åç«¯æ­£å¸¸è¿è¡Œ

æ—¶é—´å‹ç¼©è§„åˆ™ï¼š
- Excelä¸­1åˆ†é’Ÿ = å®é™…10ç§’
- æ—¶é—´å‹ç¼©æ¯” TIME_SCALE = 6
"""

import os
import sys
import time
import requests
from datetime import datetime
from decimal import Decimal
from openpyxl import load_workbook

# é…ç½®
API_BASE_URL = "http://localhost:8000/api"  # æ ¹æ®å®é™…åç«¯åœ°å€ä¿®æ”¹
TIME_SCALE = 6
TEST_INTERVAL = 10  # æ¯è¡Œæµ‹è¯•æ•°æ®é—´éš”10ç§’
DEFAULT_HEATING_TEMP = 25

# æˆ¿é—´åˆå§‹æ¸©åº¦
INITIAL_TEMPS = {
    "301": 10.0,
    "302": 15.0,
    "303": 18.0,
    "304": 12.0,
    "305": 14.0,
}

# é£é€Ÿæ˜ å°„
FAN_SPEED_MAP = {
    "é«˜": "high",
    "ä¸­": "medium",
    "ä½": "low",
}

# Excelæ–‡ä»¶è·¯å¾„
TEST_DATA_FILE = os.path.join(os.path.dirname(__file__), "data", "test_hot.xlsx")

# ============================================================
# Excelè§£æï¼ˆå¤ç”¨åŸé€»è¾‘ï¼‰
# ============================================================

def parse_test_data(filepath):
    """è§£ææµ‹è¯•æ•°æ®Excel"""
    wb = load_workbook(filepath)
    ws = wb.active
    
    rows = list(ws.iter_rows(values_only=True))
    data_rows = rows[2:]  # è·³è¿‡æ ‡é¢˜
    
    test_actions = []
    
    for row in data_rows:
        if row[0] is None or row[0] == 'è´¹ç”¨å°è®¡':
            continue
            
        time_min = row[0]
        if not isinstance(time_min, (int, float)):
            continue
        
        time_min = int(time_min)
        actions = {}
        
        for room_idx in range(5):
            room_id = f"30{room_idx + 1}"
            cell_value = row[room_idx + 1]
            
            if cell_value is not None:
                action = parse_action(cell_value)
                if action:
                    actions[room_id] = action
        
        if actions or time_min == 0:
            test_actions.append((time_min, actions))
    
    return test_actions


def parse_action(cell_value):
    """è§£æå•å…ƒæ ¼ä¸­çš„æ“ä½œæŒ‡ä»¤"""
    if cell_value is None:
        return None
    
    cell_str = str(cell_value).strip()
    
    if cell_str == "å¼€æœº":
        return {"type": "power_on"}
    elif cell_str == "å…³æœº":
        return {"type": "power_off"}
    elif cell_str in FAN_SPEED_MAP:
        return {"type": "change_speed", "fan_speed": FAN_SPEED_MAP[cell_str]}
    elif cell_str.replace(".", "").isdigit():
        return {"type": "change_temp", "target_temp": float(cell_str)}
    elif "ï¼Œ" in cell_str or "," in cell_str:
        parts = cell_str.replace("ï¼Œ", ",").split(",")
        result = {"type": "change_both"}
        for part in parts:
            part = part.strip()
            if part in FAN_SPEED_MAP:
                result["fan_speed"] = FAN_SPEED_MAP[part]
            elif part.replace(".", "").isdigit():
                result["target_temp"] = float(part)
        return result
    elif cell_str == "ä¸­å¤®ç©ºè°ƒå¯åŠ¨":
        return {"type": "system_start"}
    elif "æ£€æŸ¥ç¨‹åº" in cell_str or "è®¾ç½®" in cell_str:
        return None
    
    return None


# ============================================================
# APIå®¢æˆ·ç«¯
# ============================================================

class APIClient:
    def __init__(self, base_url):
        self.base_url = base_url.rstrip("/")
        self.session = requests.Session()
        self.session.headers.update({"Content-Type": "application/json"})
    
    def init_room(self, room_id, temp, mode="heating"):
        """åˆå§‹åŒ–æˆ¿é—´æ¸©åº¦ï¼ˆä»…DEBUGæ¨¡å¼ï¼‰"""
        url = f"{self.base_url}/admin/room/{room_id}/init/"
        response = self.session.post(url, json={"temp": temp, "mode": mode})
        return response.json()
    
    def checkin(self, room_id, customer_info=None):
        """åŠç†å…¥ä½"""
        url = f"{self.base_url}/checkin/"
        if not customer_info:
            customer_info = {
                "name": f"æµ‹è¯•é¡¾å®¢{room_id}",
                "phone": "13800138000",
                "id_card": f"1234567890{room_id}",
                "room_id": room_id,
            }
        response = self.session.post(url, json=customer_info)
        return response.json()
    
    def control_ac(self, room_id, action, **kwargs):
        """æ§åˆ¶ç©ºè°ƒ"""
        url = f"{self.base_url}/ac/control/"
        payload = {"room_id": room_id, "action": action}
        payload.update(kwargs)
        response = self.session.post(url, json=payload)
        return response.json()
    
    def get_ac_state(self, room_id):
        """è·å–ç©ºè°ƒçŠ¶æ€"""
        url = f"{self.base_url}/ac/state/{room_id}/"
        response = self.session.get(url)
        return response.json()
    
    def get_all_ac_states(self):
        """è·å–æ‰€æœ‰æˆ¿é—´çŠ¶æ€ï¼ˆç›‘æ§ï¼‰"""
        url = f"{self.base_url}/ac/monitor/"
        response = self.session.get(url)
        return response.json()
    
    def get_ac_details(self, room_id):
        """è·å–ç©ºè°ƒè¯¦å•"""
        url = f"{self.base_url}/ac/details/{room_id}/"
        response = self.session.get(url)
        return response.json()


# ============================================================
# æµ‹è¯•æ‰§è¡Œå™¨
# ============================================================

class HeatingAPITest:
    def __init__(self, api_client, test_data_file):
        self.client = api_client
        self.test_data_file = test_data_file
        self.room_ids = ["301", "302", "303", "304", "305"]
        self.room_states = {}
        self.test_start_time = None
        
        # åˆå§‹åŒ–æˆ¿é—´çŠ¶æ€ç¼“å­˜
        for room_id in self.room_ids:
            self.room_states[room_id] = {
                "target_temp": DEFAULT_HEATING_TEMP,
                "fan_speed": "medium",
                "is_on": False,
            }
    
    def setup(self):
        """æµ‹è¯•ç¯å¢ƒåˆå§‹åŒ–"""
        print("=" * 60)
        print("åˆ¶çƒ­æ¨¡å¼APIæµ‹è¯• - ç¯å¢ƒåˆå§‹åŒ–")
        print("=" * 60)
        
        # 1. ç¡®ä¿æ‰€æœ‰æˆ¿é—´å·²å…¥ä½
        print("\n1. åŠç†å…¥ä½...")
        for room_id in self.room_ids:
            result = self.client.checkin(room_id)
            if result.get("code") == 200:
                print(f"  âœ… æˆ¿é—´ {room_id} å…¥ä½æˆåŠŸ")
            else:
                print(f"  âš ï¸  æˆ¿é—´ {room_id} å…¥ä½å¤±è´¥: {result.get('message')}")
        
        # 2. åˆå§‹åŒ–æˆ¿é—´æ¸©åº¦
        print("\n2. åˆå§‹åŒ–æˆ¿é—´æ¸©åº¦...")
        for room_id, temp in INITIAL_TEMPS.items():
            result = self.client.init_room(room_id, temp, mode="heating")
            if result.get("code") == 200:
                print(f"  âœ… æˆ¿é—´ {room_id} åˆå§‹æ¸©åº¦è®¾ä¸º {temp}Â°C")
            else:
                print(f"  âš ï¸  æˆ¿é—´ {room_id} åˆå§‹åŒ–å¤±è´¥: {result.get('message')}")
        
        print("\nç¯å¢ƒåˆå§‹åŒ–å®Œæˆï¼\n")
    
    def execute_action(self, room_id, action):
        """æ‰§è¡Œå•ä¸ªæ“ä½œ"""
        action_type = action.get("type")
        
        if action_type == "power_on":
            target_temp = self.room_states[room_id]["target_temp"]
            fan_speed = "medium"  # å¼€æœºé»˜è®¤é£é€Ÿ
            
            result = self.client.control_ac(
                room_id, "power_on",
                target_temp=target_temp,
                fan_speed=fan_speed,
                mode="heating"
            )
            self.room_states[room_id]["is_on"] = True
            self.room_states[room_id]["fan_speed"] = fan_speed
            print(f"    ğŸ”› å¼€æœº (ç›®æ ‡{target_temp}Â°C, {fan_speed})")
            
        elif action_type == "power_off":
            result = self.client.control_ac(room_id, "power_off")
            self.room_states[room_id]["is_on"] = False
            print(f"    â¹ï¸  å…³æœº")
            
        elif action_type == "change_temp":
            target_temp = action.get("target_temp")
            self.room_states[room_id]["target_temp"] = target_temp
            
            # è°ƒæ¸©è¯·æ±‚ç›´æ¥å‘é€ï¼ˆé˜²æŠ–é€»è¾‘ï¼‰
            result = self.client.control_ac(
                room_id, "change_temp",
                target_temp=target_temp,
                mode="heating"
            )
            print(f"    ğŸŒ¡ï¸  è°ƒæ¸© -> {target_temp}Â°C")
            
        elif action_type == "change_speed":
            fan_speed = action.get("fan_speed")
            self.room_states[room_id]["fan_speed"] = fan_speed
            
            result = self.client.control_ac(
                room_id, "change_speed",
                fan_speed=fan_speed
            )
            print(f"    ğŸ’¨ è°ƒé£é€Ÿ -> {fan_speed}")
            
        elif action_type == "change_both":
            target_temp = action.get("target_temp")
            fan_speed = action.get("fan_speed")
            
            if target_temp:
                self.room_states[room_id]["target_temp"] = target_temp
                result = self.client.control_ac(
                    room_id, "change_temp",
                    target_temp=target_temp,
                    mode="heating"
                )
            
            if fan_speed:
                self.room_states[room_id]["fan_speed"] = fan_speed
                result = self.client.control_ac(
                    room_id, "change_speed",
                    fan_speed=fan_speed
                )
            
            print(f"    ğŸ”„ è°ƒæ¸©={target_temp}Â°C, è°ƒé£é€Ÿ={fan_speed}")
        
        # æ£€æŸ¥è¯·æ±‚æ˜¯å¦æˆåŠŸ
        if result.get("code") != 200:
            print(f"    âŒ å¤±è´¥: {result.get('message')}")
        else:
            # å¯¹äºpower_on/change_speedï¼Œæ›´æ–°çŠ¶æ€
            if action_type in ["power_on", "change_speed"]:
                self.room_states[room_id].update(result.get("data", {}))
    
    def print_status(self, time_min):
        """æ‰“å°å½“å‰æ‰€æœ‰æˆ¿é—´çŠ¶æ€"""
        print(f"\n  ğŸ“Š [çŠ¶æ€] æ—¶é—´={time_min}åˆ†é’Ÿ")
        print("  " + "-" * 90)
        print(f"  {'æˆ¿é—´':<8} {'çŠ¶æ€':<12} {'å½“å‰æ¸©åº¦':<10} {'ç›®æ ‡æ¸©åº¦':<10} {'é£é€Ÿ':<8} {'è´¹ç”¨':<8} {'é˜Ÿåˆ—':<10}")
        print("  " + "-" * 90)
        
        all_states = self.client.get_all_ac_states().get("data", [])
        state_map = {s["room_id"]: s for s in all_states}
        
        for room_id in self.room_ids:
            state = state_map.get(room_id, {})
            status = state.get("status", "off")
            current = state.get("current_temp", 0)
            target = state.get("target_temp", 0)
            fan_speed = state.get("fan_speed", "-")
            cost = state.get("cost", 0)
            
            # æ ‡è®°é˜Ÿåˆ—ä½ç½®
            queue_info = ""
            if status == "on":
                queue_info = "[æœåŠ¡]"
            elif status == "waiting":
                remaining = state.get("remaining_wait", 0)
                queue_info = f"[ç­‰{remaining:.0f}s]"
            
            print(f"  {room_id:<8} {status:<12} {current:<10.1f} {target:<10.1f} {fan_speed:<8} {cost:<8.2f} {queue_info:<10}")
        
        print("  " + "-" * 90)
    
    def print_final_report(self):
        """æ‰“å°æœ€ç»ˆæŠ¥å‘Š"""
        print("\n" + "=" * 60)
        print("æµ‹è¯•å®Œæˆ - æœ€ç»ˆæŠ¥å‘Š")
        print("=" * 60)
        
        total_cost = Decimal("0.00")
        total_energy = 0.0
        
        print("\nğŸ’° è´¹ç”¨æ±‡æ€»ï¼ˆä»è¯¦å•è®°å½•ç»Ÿè®¡ï¼‰:")
        print("-" * 60)
        
        for room_id in self.room_ids:
            details = self.client.get_ac_details(room_id).get("data", {})
            summary = details.get("summary", {})
            room_cost = Decimal(str(summary.get("total_cost", 0)))
            room_energy = float(summary.get("total_energy", 0))
            
            total_cost += room_cost
            total_energy += room_energy
            
            print(f"  æˆ¿é—´ {room_id}: è´¹ç”¨={room_cost:.2f}å…ƒ, èƒ½è€—={room_energy:.2f}åº¦")
        
        print("-" * 60)
        print(f"  æ€»è®¡: è´¹ç”¨={total_cost:.2f}å…ƒ, èƒ½è€—={total_energy:.2f}åº¦")
        
        # æ‰“å°æ¯ä¸ªæˆ¿é—´çš„è¯¦ç»†è®°å½•
        print("\n" + "=" * 100)
        print("å„æˆ¿é—´è¯¦ç»†ç©ºè°ƒè®°å½•")
        print("=" * 100)
        
        for room_id in self.room_ids:
            print(f"\nğŸ“‹ æˆ¿é—´ {room_id} è¯¦å•:")
            details = self.client.get_ac_details(room_id).get("data", {})
            records = details.get("details", [])
            
            if not records:
                print("  (æ— è®°å½•)")
                continue
            
            print(f"  {'åºå·':<4} {'å¼€å§‹æ—¶é—´':<20} {'æ—¶é•¿(ç§’)':<10} {'èµ·å§‹æ¸©åº¦':<10} {'ç›®æ ‡æ¸©åº¦':<10} {'é£é€Ÿ':<6} {'èƒ½è€—':<8} {'è´¹ç”¨':<8}")
            print("  " + "-" * 90)
            
            for r in records:
                print(f"  {r['seq']:<4} {r['start_time']:<20} {r['duration_seconds']:<10} "
                      f"{r['start_temp']:<10.1f} {r['target_temp']:<10.1f} {r['fan_speed']:<6} "
                      f"{r['energy']:<8.2f} {r['cost']:<8.2f}")
    
    def run_test(self):
        """è¿è¡Œæµ‹è¯•"""
        # è§£ææµ‹è¯•æ•°æ®
        print(f"ğŸ“‚ åŠ è½½æµ‹è¯•æ•°æ®: {self.test_data_file}")
        test_data = parse_test_data(self.test_data_file)
        print(f"âœ… å…±è§£æ {len(test_data)} ä¸ªæ—¶é—´ç‚¹\n")
        
        # åˆå§‹åŒ–
        self.setup()
        
        print("=" * 60)
        print("å¼€å§‹æ‰§è¡Œæµ‹è¯•")
        print(f"â±ï¸  æ—¶é—´å‹ç¼©æ¯”: {TIME_SCALE}x (10ç§’æµ‹è¯• = 60ç§’ç³»ç»Ÿæ—¶é—´)")
        print("=" * 60)
        
        self.test_start_time = time.time()
        
        for time_min, actions in test_data:
            # ç­‰å¾…åˆ°æŒ‡å®šæ—¶é—´ç‚¹
            target_test_time = time_min * TEST_INTERVAL
            current_test_time = time.time() - self.test_start_time
            
            if target_test_time > current_test_time:
                wait_time = target_test_time - current_test_time
                print(f"\nâ³ ç­‰å¾… {wait_time:.1f} ç§’åˆ°è¾¾æ—¶é—´ç‚¹ {time_min} åˆ†é’Ÿ...")
                time.sleep(wait_time)
            
            print(f"\n{'='*60}")
            print(f"â° æ—¶é—´ç‚¹: {time_min} åˆ†é’Ÿ (å·²è¿è¡Œ: {current_test_time:.1f}ç§’)")
            print(f"{'='*60}")
            
            if time_min == 0:
                print("  ğŸ¬ ç³»ç»Ÿå¯åŠ¨ï¼Œè®¾ç½®åˆ¶çƒ­æ¨¡å¼")
                continue
            
            # æ‰§è¡Œæ“ä½œ
            if actions:
                print("  ğŸ“ æ‰§è¡Œæ“ä½œ:")
                for room_id, action in actions.items():
                    print(f"    [{room_id}] ", end="")
                    self.execute_action(room_id, action)
            else:
                print("  (æ— æ“ä½œ)")
            
            # æ‰“å°çŠ¶æ€
            self.print_status(time_min)
        
        # æµ‹è¯•ç»“æŸ
        self.print_final_report()
        
        print("\nâœ… æµ‹è¯•æ‰§è¡Œå®Œæ¯•ï¼")


# ============================================================
# ä¸»å‡½æ•°
# ============================================================

def main():
    """ä¸»å‡½æ•°"""
    print("=" * 60)
    print("åˆ¶çƒ­æ¨¡å¼APIæµ‹è¯•é©±åŠ¨å™¨")
    print(f"å¯åŠ¨æ—¶é—´: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print(f"åç«¯åœ°å€: {API_BASE_URL}")
    print("=" * 60)
    
    # æ£€æŸ¥åç«¯æ˜¯å¦å¯è®¿é—®
    try:
        response = requests.get(f"{API_BASE_URL}/rooms/", timeout=5)
        if response.status_code != 200:
            print("âŒ åç«¯APIæ— æ³•è®¿é—®ï¼Œè¯·ç¡®ä¿DjangoæœåŠ¡æ­£åœ¨è¿è¡Œ")
            sys.exit(1)
        print("âœ… åç«¯APIè¿æ¥æ­£å¸¸\n")
    except requests.exceptions.RequestException as e:
        print(f"âŒ æ— æ³•è¿æ¥åç«¯: {e}")
        sys.exit(1)
    
    # åˆ›å»ºå®¢æˆ·ç«¯å’Œæµ‹è¯•å®ä¾‹
    client = APIClient(API_BASE_URL)
    test = HeatingAPITest(client, TEST_DATA_FILE)
    
    try:
        test.run_test()
    except KeyboardInterrupt:
        print("\n\nâš ï¸  æµ‹è¯•è¢«ç”¨æˆ·ä¸­æ–­")
    except Exception as e:
        print(f"\n\nâŒ æµ‹è¯•æ‰§è¡Œå‡ºé”™: {e}")
        import traceback
        traceback.print_exc()


if __name__ == "__main__":
    main()